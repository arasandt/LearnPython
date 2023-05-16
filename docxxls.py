# import argparse
# import re
# import xml.etree.ElementTree as ET
# import zipfile
# import os
# import sys


# nsmap = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


# def qn(tag):
#     """
#     Stands for 'qualified name', a utility function to turn a namespace
#     prefixed tag name into a Clark-notation qualified tag name for lxml. For
#     example, ``qn('p:cSld')`` returns ``'{http://schemas.../main}cSld'``.
#     Source: https://github.com/python-openxml/python-docx/
#     """
#     prefix, tagroot = tag.split(":")
#     uri = nsmap[prefix]
#     return "{{{}}}{}".format(uri, tagroot)


# def xml2text(xml):
#     """
#     A string representing the textual content of this run, with content
#     child elements like ``<w:tab/>`` translated to their Python
#     equivalent.
#     Adapted from: https://github.com/python-openxml/python-docx/
#     """
#     text = ""
#     root = ET.fromstring(xml)
#     for child in root.iter():
#         if child.tag == qn("w:t"):
#             t_text = child.text
#             text += t_text if t_text is not None else ""
#         elif child.tag == qn("w:tab"):
#             text += "\t"
#         elif child.tag in (qn("w:br"), qn("w:cr")):
#             text += "\n"
#         elif child.tag == qn("w:p"):
#             text += "\n\n"
#     return text


# def process(docx, img_dir=None):
#     text = ""

#     # unzip the docx in memory
#     zipf = zipfile.ZipFile(docx)
#     filelist = zipf.namelist()

#     # get header text
#     # there can be 3 header files in the zip
#     header_xmls = "word/header[0-9]*.xml"
#     for fname in filelist:
#         if re.match(header_xmls, fname):
#             text += xml2text(zipf.read(fname))

#     # get main text
#     doc_xml = "word/document.xml"
#     text += xml2text(zipf.read(doc_xml))

#     # get footer text
#     # there can be 3 footer files in the zip
#     footer_xmls = "word/footer[0-9]*.xml"
#     for fname in filelist:
#         if re.match(footer_xmls, fname):
#             text += xml2text(zipf.read(fname))

#     if img_dir is not None:
#         # extract images
#         for fname in filelist:
#             _, extension = os.path.splitext(fname)
#             if extension in [".jpg", ".jpeg", ".png", ".bmp"]:
#                 dst_fname = os.path.join(img_dir, os.path.basename(fname))
#                 with open(dst_fname, "wb") as dst_f:
#                     dst_f.write(zipf.read(fname))

#     zipf.close()
#     return text.strip()


# text = process("Linux Install Guide.docx")
# text = re.sub(r"\s", " ", text)
# print(text)

from openpyxl import load_workbook
import openpyxl


def process(filename):
    wb_obj = load_workbook(filename=filename)

    excel_data = ""
    for sheet_name in wb_obj.sheetnames:
        print(f"Sheet Name: {sheet_name}")
        sheet = wb_obj[sheet_name]
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Skip this cell
                    continue
                # print(f"{cell.value}")
                if cell.value is not None:
                    excel_data += str(cell.value) + " "
    return excel_data


text = process("Distinguished Masters.xlsx")
print(text)

# for i in range(1, max_col + 1):

#     cell_obj = sheet_obj.cell(row=2, column=i)
#     print(cell_obj.value, end=" ")
